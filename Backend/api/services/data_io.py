"""
Reusable Import & Export core services for WealthWise.

This module is intentionally domain-independent. It exposes a clean, extensible
API (``FileParser``, ``MappingService``, ``ValidationService``,
``ImportService``, ``ExportService``) that any feature module can reuse for
importing or exporting tabular data without knowing about the finance domain.

The actual persistence of parsed rows is delegated to an ``ImportSink`` provided
by the consuming module (e.g. the transaction importer), keeping this core free
of any model imports.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

DATE_FORMATS = (
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%d %b %Y",
    "%d %B %Y",
    "%b %d, %Y",
    "%B %d, %Y",
    "%d.%m.%Y",
    "%Y/%m/%d",
    "%Y%m%d",
)


class ParseError(Exception):
    """Raised when a file cannot be parsed."""


@dataclass
class ParsedRow:
    """A single parsed row with its raw and normalized representations."""

    index: int
    raw: Dict[str, Any]
    normalized: Dict[str, Any] = field(default_factory=dict)
    errors: List[str] = field(default_factory=list)
    valid: bool = True
    skipped: bool = False


@dataclass
class ParseResult:
    """Result of parsing an uploaded file into normalized rows."""

    headers: List[str]
    rows: List[ParsedRow]
    sheet_names: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def valid_rows(self) -> List[ParsedRow]:
        return [r for r in self.rows if r.valid and not r.skipped]

    @property
    def invalid_rows(self) -> List[ParsedRow]:
        return [r for r in self.rows if not r.valid]

    @property
    def total(self) -> int:
        return len(self.rows)


class FileParser:
    """Parses CSV / Excel / PDF files into a normalized, domain-agnostic shape.

    The parser does not understand business meaning; it simply extracts tabular
    rows (and, for PDF bank statements, attempts to extract transaction-like
    lines). Meaning is applied later via the ``MappingService``.
    """

    MAX_FILE_BYTES = 25 * 1024 * 1024  # 25 MB
    ALLOWED_EXTENSIONS = {".csv", ".xls", ".xlsx", ".pdf"}

    @classmethod
    def detect_format(cls, filename: str) -> str:
        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in cls.ALLOWED_EXTENSIONS:
            raise ParseError(
                f"Unsupported file type '{ext or filename}'. "
                f"Allowed: {', '.join(sorted(cls.ALLOWED_EXTENSIONS))}."
            )
        return ext.lstrip(".")

    @classmethod
    def validate_file(cls, filename: str, size: int) -> None:
        if size > cls.MAX_FILE_BYTES:
            raise ParseError(
                f"File too large ({size // 1024} KB). Maximum allowed is "
                f"{cls.MAX_FILE_BYTES // 1024} KB."
            )
        cls.detect_format(filename)

    def parse(self, filename: str, content: bytes) -> ParseResult:
        fmt = self.detect_format(filename)
        if fmt == "csv":
            return self._parse_csv(content)
        if fmt in ("xls", "xlsx"):
            return self._parse_excel(content)
        if fmt == "pdf":
            return self._parse_pdf(content)
        raise ParseError(f"Unsupported format: {fmt}")

    # ------------------------------------------------------------------ CSV
    def _parse_csv(self, content: bytes) -> ParseResult:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        reader = csv.reader(io.StringIO(text))
        matrix = [row for row in reader if any(cell.strip() for cell in row)]
        if not matrix:
            raise ParseError("The file appears to be empty.")

        headers = [h.strip() for h in matrix[0]]
        rows: List[ParsedRow] = []
        for i, raw in enumerate(matrix[1:], start=0):
            record = {headers[j]: (raw[j] if j < len(raw) else "") for j in range(len(headers))}
            rows.append(ParsedRow(index=i, raw=record))

        warnings = self._detect_empty_rows_warning(rows)
        return ParseResult(headers=headers, rows=rows, warnings=warnings)

    # ---------------------------------------------------------------- Excel
    def _parse_excel(self, content: bytes) -> ParseResult:
        try:
            import openpyxl  # noqa: F401  (imported lazily for clarity)
        except ImportError as exc:  # pragma: no cover - environment guard
            raise ParseError(
                "Excel support requires 'openpyxl'. Install it to import .xls/.xlsx files."
            ) from exc

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        target = sheet_names[0]
        sheet = workbook[target]

        matrix: List[List[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            matrix.append([("" if cell is None else str(cell)) for cell in row])
        # Drop fully empty trailing rows.
        matrix = [r for r in matrix if any(cell.strip() for cell in r)]
        if not matrix:
            raise ParseError("The selected worksheet appears to be empty.")

        headers = [str(h).strip() for h in matrix[0]]
        rows: List[ParsedRow] = []
        for i, raw in enumerate(matrix[1:], start=0):
            record = {headers[j]: (raw[j] if j < len(raw) else "") for j in range(len(headers))}
            rows.append(ParsedRow(index=i, raw=record))

        warnings = self._detect_empty_rows_warning(rows)
        return ParseResult(
            headers=headers, rows=rows, sheet_names=sheet_names, warnings=warnings
        )

    # ------------------------------------------------------------------ PDF
    def _parse_pdf(self, content: bytes) -> ParseResult:
        try:
            import pdfplumber  # type: ignore
        except ImportError as exc:  # pragma: no cover - environment guard
            raise ParseError(
                "PDF bank-statement parsing requires 'pdfplumber'. "
                "Install it to import PDF files."
            ) from exc

        table_rows: List[List[str]] = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                for tbl in page.extract_tables():
                    for r in tbl:
                        table_rows.append([("" if c is None else str(c).strip()) for c in r])
                # Fallback: if the page produced no real table, or every table row
                # collapsed into a single cell, rebuild rows from raw text lines.
                if not table_rows or all(len(r) <= 1 for r in table_rows):
                    table_rows = FileParser._extract_statement_lines(page.extract_text() or "")

        if not table_rows:
            raise ParseError("Could not extract any tabular data from the PDF.")

        # Degenerate tables (everything in one column) need whitespace splitting.
        if table_rows and all(len(r) <= 1 for r in table_rows):
            table_rows = [self._split_statement_line(r[0]) for r in table_rows if r]

        # Treat the first non-trivial row as the header if it looks like one.
        headers, data_rows = self._infer_pdf_header(table_rows)
        rows: List[ParsedRow] = []
        for i, raw in enumerate(data_rows, start=0):
            record = {headers[j]: (raw[j] if j < len(raw) else "") for j in range(len(headers))}
            rows.append(ParsedRow(index=i, raw=record))

        warnings = ["PDF parsing is best-effort; please review extracted rows."]
        warnings += self._detect_empty_rows_warning(rows)
        return ParseResult(headers=headers, rows=rows, warnings=warnings)

    # --------------------------------------------------------------- helpers
    @staticmethod
    def _split_statement_line(line: str) -> List[str]:
        """Split a single-column PDF statement line into columns.

        Bank statements are usually monospaced, so we first split on runs of two
        or more spaces / tabs. If that still leaves a single chunk (e.g. the
        PDF used single spaces between cells), fall back to regex tokenization
        that isolates the leading date and trailing numeric amounts and treats
        whatever remains as the description.
        """
        cols = [c.strip() for c in re.split(r"\s{2,}|\t", line) if c.strip()]
        if len(cols) > 1:
            return cols

        date_m = re.search(r"(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})", line)
        # Numeric tokens (amounts/balance), keeping order and positions.
        num_iter = list(re.finditer(r"[-+]?[\d,]+\.?\d*", line))
        parts: List[str] = []
        if date_m:
            parts.append(date_m.group(1))
            rest = line[date_m.end():]
        else:
            rest = line
        # Description = text before the first numeric token after the date.
        if num_iter:
            first_num = num_iter[0]
            desc = rest[: first_num.start()].strip()
            parts.append(desc)
            for nm in num_iter:
                parts.append(nm.group(0).strip())
        else:
            parts.append(rest.strip())
        return [p for p in parts if p]

    @staticmethod
    def _extract_statement_lines(text: str) -> List[List[str]]:
        """Fallback line extraction for PDFs without detectable tables.

        Looks for lines containing a date and a numeric amount and splits them
        into columns.
        """
        out: List[List[str]] = []
        date_re = re.compile(r"(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})")
        amount_re = re.compile(r"[-+]?\d[\d,]*\.?\d*")
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            if date_re.search(line) and amount_re.search(line):
                out.append(FileParser._split_statement_line(line))
        return out

    @staticmethod
    def _infer_pdf_header(rows: List[List[str]]) -> Tuple[List[str], List[List[str]]]:
        header_candidate = rows[0]
        looks_like_header = any(
            kw in " ".join(header_candidate).lower()
            for kw in ("date", "desc", "particular", "debit", "credit", "amount", "balance", "narration")
        )
        if looks_like_header:
            return [h.strip() for h in header_candidate], rows[1:]
        # No header: synthesize generic column names.
        width = max(len(r) for r in rows)
        return [f"column_{i + 1}" for i in range(width)], rows

    @staticmethod
    def _detect_empty_rows_warning(rows: List[ParsedRow]) -> List[str]:
        empties = sum(1 for r in rows if not any(str(v).strip() for v in r.raw.values()))
        if empties:
            return [f"{empties} empty or near-empty row(s) were detected."]
        return []


# Standard field keys any importer can map source columns to.
STANDARD_FIELDS = [
    "date",
    "description",
    "merchant",
    "debit",
    "credit",
    "amount",
    "currency",
    "balance",
    "reference",
]


class MappingService:
    """Maps arbitrary source column names onto standard field keys.

    Supports fuzzy auto-detection so users rarely have to map columns by hand,
    and persists named templates for reuse.
    """

    _ALIASES = {
        "date": ["date", "txn date", "transaction date", "posted", "value date", "time"],
        "description": ["description", "desc", "particulars", "particular", "narration", "details", "remark", "remarks"],
        "merchant": ["merchant", "payee", "beneficiary", "party", "name"],
        "debit": ["debit", "withdrawal", "withdraw", "dr", "expense", "paid", "spent"],
        "credit": ["credit", "deposit", "cr", "income", "received", "refund"],
        "amount": ["amount", "amt", "value", "sum", "total"],
        "currency": ["currency", "ccy", "cur"],
        "balance": ["balance", "running balance", "closing balance", "bal"],
        "reference": ["reference", "ref", "ref no", "ref number", "cheque", "chq", "txn id", "transaction id"],
    }

    @classmethod
    def auto_detect(cls, headers: Sequence[str]) -> Dict[str, str]:
        """Return a mapping of ``standard_field -> source_header``."""
        mapping: Dict[str, str] = {}
        normalized = {h.lower().strip(): h for h in headers}
        for standard, aliases in cls._ALIASES.items():
            for alias in aliases:
                for key, original in normalized.items():
                    if alias == key or key.startswith(alias) or key.endswith(alias):
                        if original not in mapping.values():
                            mapping[standard] = original
                        break
                if standard in mapping:
                    break
        return mapping

    @classmethod
    def normalize_row(
        cls, row: ParsedRow, mapping: Dict[str, str]
    ) -> Dict[str, Any]:
        """Apply a column mapping to a single parsed row."""
        out: Dict[str, Any] = {}
        for standard, source in mapping.items():
            out[standard] = row.raw.get(source, "")
        return out


class ValidationService:
    """Validates and normalizes mapped rows against a set of field rules."""

    REQUIRED_FIELDS = ["date", "amount"]

    @staticmethod
    def normalize_date(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        # Handle datetime strings.
        text = text.replace("T", " ").split(" ")[0] if "T" in text else text
        for fmt in DATE_FORMATS:
            try:
                from datetime import datetime

                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        # Last resort: let dateutil-style parsing try, else None.
        return None

    @staticmethod
    def normalize_amount(value: Any) -> Optional[Decimal]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        # Strip currency symbols, thousands separators, spaces, parentheses.
        text = text.replace(",", "").replace("₹", "").replace("$", "")
        text = text.replace("€", "").replace("£", "").replace("¥", "").strip()
        negative = text.startswith("(") and text.endswith(")")
        text = text.strip("()")
        try:
            amount = Decimal(text)
        except (InvalidOperation, ValueError):
            return None
        return -amount if negative else amount

    def validate(self, rows: List[ParsedRow], mapping: Dict[str, str]) -> List[ParsedRow]:
        seen: Dict[Tuple[str, str, str], int] = {}
        for row in rows:
            normalized = MappingService.normalize_row(row, mapping)
            errors: List[str] = []

            date = self.normalize_date(normalized.get("date"))
            if date is None and "date" in self.REQUIRED_FIELDS:
                errors.append("Missing or unrecognized date.")
            normalized["date"] = date

            amount = self.normalize_amount(normalized.get("amount"))
            debit = self.normalize_amount(normalized.get("debit"))
            credit = self.normalize_amount(normalized.get("credit"))
            if amount is None:
                if debit is not None:
                    amount = -abs(debit)
                elif credit is not None:
                    amount = abs(credit)
            if amount is None and "amount" in self.REQUIRED_FIELDS:
                errors.append("Missing or invalid amount.")
            normalized["amount"] = amount

            normalized["description"] = str(normalized.get("description") or "").strip()
            normalized["merchant"] = str(normalized.get("merchant") or "").strip()
            normalized["currency"] = str(normalized.get("currency") or "").strip().upper() or None
            normalized["balance"] = self.normalize_amount(normalized.get("balance"))
            normalized["reference"] = str(normalized.get("reference") or "").strip() or None

            # Duplicate detection on (date, description, amount).
            key = (
                str(normalized.get("date")),
                normalized.get("description", ""),
                str(normalized.get("amount")),
            )
            if key in seen:
                errors.append("Duplicate of an earlier row.")
            else:
                seen[key] = 1

            row.normalized = normalized
            row.errors = errors
            row.valid = len(errors) == 0
        return rows


class ImportService:
    """Orchestrates parse -> map -> validate -> persist.

    ``sink`` is a callable that receives the list of valid normalized records
    and returns the number of records created. This keeps the service free of
    any concrete model, so any module can plug in its own persistence.
    """

    def __init__(self, parser: Optional[FileParser] = None) -> None:
        self.parser = parser or FileParser()

    def parse_file(self, filename: str, content: bytes) -> ParseResult:
        return self.parser.parse(filename, content)

    def analyze(
        self, result: ParseResult, mapping: Optional[Dict[str, str]] = None
    ) -> ParseResult:
        if mapping is None:
            mapping = MappingService.auto_detect(result.headers)
        self._last_mapping = mapping
        ValidationService().validate(result.rows, mapping)
        return result

    def commit(
        self,
        result: ParseResult,
        sink: Callable[[List[Dict[str, Any]]], int],
        skip_invalid: bool = True,
    ) -> Dict[str, int]:
        rows = result.valid_rows if skip_invalid else [r for r in result.rows if r.valid]
        count = sink([r.normalized for r in rows])
        return {
            "imported": count,
            "skipped": result.total - len(rows),
            "total": result.total,
        }


class ExportService:
    """Generates exports in multiple formats from a list of records.

    Records are plain dicts keyed by column name so any module can export its
    data without coupling to a specific model.
    """

    @staticmethod
    def to_csv(columns: Sequence[str], records: Iterable[Dict[str, Any]]) -> str:
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=list(columns), extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            writer.writerow({c: rec.get(c, "") for c in columns})
        return buffer.getvalue()

    @staticmethod
    def to_json(columns: Sequence[str], records: Iterable[Dict[str, Any]]) -> str:
        import json

        cleaned = [{c: rec.get(c, "") for c in columns} for rec in records]
        return json.dumps(cleaned, indent=2, default=str)

    @staticmethod
    def to_excel(columns: Sequence[str], records: Iterable[Dict[str, Any]], sheet_name: str = "Export") -> bytes:
        try:
            from openpyxl import Workbook
        except ImportError as exc:  # pragma: no cover - environment guard
            raise ParseError("Excel export requires 'openpyxl'.") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        ws.append(list(columns))
        for rec in records:
            ws.append([rec.get(c, "") for c in columns])
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def to_pdf_table(title: str, columns: Sequence[str], records: Iterable[Dict[str, Any]]) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        data = [[Paragraph(f"<b>{c}</b>", getSampleStyleSheet()["Normal"]) for c in columns]]
        for rec in records:
            data.append([str(rec.get(c, "")) for c in columns])

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = [Paragraph(title, getSampleStyleSheet()["Title"]), Spacer(1, 12)]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()

    @classmethod
    def render(
        cls,
        fmt: str,
        title: str,
        columns: Sequence[str],
        records: Iterable[Dict[str, Any]],
    ) -> Tuple[bytes, str, str]:
        fmt = fmt.lower()
        if fmt == "csv":
            body = cls.to_csv(columns, records).encode("utf-8")
            return body, "text/csv", "csv"
        if fmt == "json":
            body = cls.to_json(columns, records).encode("utf-8")
            return body, "application/json", "json"
        if fmt in ("xlsx", "xls", "excel"):
            return cls.to_excel(columns, records), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
        if fmt == "pdf":
            return cls.to_pdf_table(title, columns, records), "application/pdf", "pdf"
        raise ParseError(f"Unsupported export format: {fmt}")
